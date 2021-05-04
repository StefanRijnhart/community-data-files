# Copyright 2021 Stefan Rijnhart <stefan@opener.amsterdam>
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl)
# Transforms the spreadsheet found at https://www.cepa.be/wp-content/uploads\
#    /ADR_2019_BijlageA_deel3_Tabel_A_EXCEL_FORMAAT.xlsx
# to adr.goods XML data.

import re
import sys

from lxml import etree
from openpyxl import load_workbook  # pylint: disable=W7936

skiprows = 3
columns = {
    0: "un_number",
    2: "name",
    5: "class_id",
    6: "classification_code",
    8: "label_ids",
    20: "transport_category",
}

# Articles containing dangerous goods
# (as opposed to dangerous goods themselves)
article_labels = {
    "3537": "2.1",
    "3538": "2.2",
    "3539": "2.3",
    "3540": "3",
    "3541": "4.1",
    "3542": "4.2",
    "3543": "4.3",
    "3544": "5.1",
    "3545": "5.2",
    "3546": "6.1",
    "3547": "8",
    "3548": "9",
}

transformers = {}


def transformer(func):
    """
    Decorator to add functions to the `transformers` dictionary

    Functions are added by their function name
    """
    transformers[func.__name__] = func
    return func


@transformer
def un_number(record, value, row):
    record.attrib.update(
        {
            "id": "adr_goods_%s" % value,
            "model": "adr.goods",
        }
    )
    etree.SubElement(record, "field", attrib={"name": "un_number"}).text = value


@transformer
def name(record, value, row):
    value = value.strip().replace("\n", "")
    etree.SubElement(record, "field", attrib={"name": "name"}).text = value


@transformer
def class_id(record, value, row):
    etree.SubElement(
        record,
        "field",
        attrib={
            "name": "class_id",
            "ref": "adr_class_%s" % value.replace(".", "_"),
        },
    )


@transformer
def classification_code(record, value, row):
    if value is None:  # Not defined for code 0190, for example
        return
    # In case multiple codes are possible, take the first (rare)
    code = value.split(" of ")[0]
    code = code.replace(",", ".")  # Typo in the sheet
    etree.SubElement(
        record, "field", attrib={"name": "classification_code"}
    ).text = code


@transformer
def transport_category(record, value, row):
    valid_categories = [
        "0",
        "1",
        "2",
        "3",
        "4",
        "-",
        "CARRIAGE_PROHIBITED",
        "NOT_SUBJECT_TO_ADR",
    ]
    valid_tunnel_codes = [
        "B",
        "B1000C",
        "B/D",
        "B/E",
        "C5000D",
        "C",
        "C/D",
        "C/E",
        "D",
        "D/E",
        "E",
        "-",
        "CARRIAGE_PROHIBITED",
        "NOT_SUBJECT_TO_ADR",
    ]
    if value is None:
        if "VERVOER VERBODEN" in str(row):
            # codes 0020, 0021
            category = "CARRIAGE_PROHIBITED"
            tunnel_restriction_code = "CARRIAGE_PROHIBITED"
        elif "NIET ONDERWORPEN AAN HET ADR" in str(row):
            category = "NOT_SUBJECT_TO_ADR"
            tunnel_restriction_code = "NOT_SUBJECT_TO_ADR"
        elif str(row[0]) in ["2071", "3363"]:  # known exceptions
            category = "-"
            tunnel_restriction_code = "-"
    else:
        match = re.search(r"(.*)\(([^\)]+)\)", value, re.DOTALL)
        if not match:
            raise ValueError(
                "Unknown value for transport code/tunnel restriction code: "
                "%s" % value,
                row,
            )

        category = match.groups()[0].strip()
        tunnel_restriction_code = match.groups()[1].strip()
    if "BP671" in category:
        # Special provision 671. Category depending on packing group
        # or else "2"
        category = "2"
    if category == "_":
        category = "-"
    if category not in valid_categories:
        raise ValueError(
            "Invalid transport category {} in cell value {}".format(category, value)
        )
    if tunnel_restriction_code not in valid_tunnel_codes:
        raise ValueError(
            "Invalid tunnel restriction code %s in cell value %s"
            % (tunnel_restriction_code, value)
        )
    etree.SubElement(
        record,
        "field",
        attrib={
            "name": "transport_category",
        },
    ).text = category
    etree.SubElement(
        record,
        "field",
        attrib={
            "name": "tunnel_restriction_code",
        },
    ).text = tunnel_restriction_code


@transformer
def label_ids(record, value, row):
    valid_labels = [
        "1",
        "1.4",
        "1.5",
        "1.6",
        "2.1",
        "2.2",
        "2.3",
        "3",
        "4.1",
        "4.2",
        "4.3",
        "5.1",
        "5.2",
        "6.1",
        "6.2",
        "7A",
        "7B",
        "7C",
        "7E",
        "8",
        "9",
        "9A",
    ]
    labels = [label.strip() for label in (value or "").split("+")]
    if "7X" in labels:
        labels.remove("7X")
        # Pick any one
        labels += ["7A", "7B", "7C", "7E"]
    label_refs = []
    for label in labels:
        if not label or label == "GEEN":  # Meaning: none
            continue
        if any(
            val in label
            for val in (
                # Carriage prohibited
                "VERVOER VERBODEN",
                # Not subject to ADR
                "NIET ONDERWORPEN AAN HET ADR",
            )
        ):
            break
        if "5.2.2.1.12" in label:
            if str(row[0]) in article_labels:
                label = article_labels[str(row[0])]
        if label not in valid_labels:
            raise ValueError("Invalid label {} in cell value {}".format(label, value))
        label_refs.append("ref('adr_label_%s')" % label.replace(".", "_"))
    expression = "[(6, 0, [%s])]" % ", ".join(label_refs)
    etree.SubElement(
        record,
        "field",
        attrib={
            "name": "label_ids",
            "eval": expression,
        },
    )


def transform_row(root, row):
    record = etree.SubElement(root, "record")
    for index, field in columns.items():
        value = row[index]
        if isinstance(value, (int, float)):
            value = str(value)
        try:
            transformers[field](record, value, row)
        except (ValueError, AttributeError) as e:
            raise ValueError("Could not transform row {}: {}".format(row, e))


def import_adr_multilang_xlsx(argv):
    root = etree.Element("odoo")
    sheet = load_workbook(argv[0]).active
    count = 0
    seen = set()
    for row in sheet.iter_rows(values_only=True):
        count += 1
        if count <= skiprows:
            continue
        if row[0] is None:  # Emtpy rows
            continue
        code = str(row[0]).strip()
        if code in seen:
            continue
        seen.add(code)
        transform_row(root, row)
    print(  # pylint: disable=W8116
        etree.tostring(
            root, pretty_print=True, xml_declaration=True, encoding="utf-8"
        ).decode("utf-8")
    ),


if __name__ == "__main__":
    import_adr_multilang_xlsx(sys.argv[1:])
